#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Copyright 2023 Jetperch LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Perform long-term current monitoring.

CAUTION: UNDER DEVELOPMENT
See https://forum.joulescope.com/t/exported-jls-data-to-an-xlsx-file/580

Inspect current data for errant conditions.  On an errant condition,
record data for analysis.  For now, records 50 Hz statistics
data to both a JLS v2 file and an Excel XLSX file.

See https://github.com/jetperch/jls
See https://github.com/jetperch/joulescope_driver
"""

from pyjls import Writer, SignalType, DataType
from pyjoulescope_driver import Driver
import openpyxl
import time
import numpy as np
import datetime


# todo should probably be part of pyjoulescope_driver
EPOCH = datetime.datetime(2018, 1, 1, tzinfo=datetime.timezone.utc).timestamp()  # seconds
SECOND = (1 << 30)


def _time64_to_datetime(t64) -> datetime.datetime:
    ts = (t64 / SECOND) + EPOCH
    return datetime.datetime.fromtimestamp(ts)


def _time64_to_filename(t64, suffix=None) -> str:
    suffix = '.jls' if suffix is None else str(suffix)
    dt = _time64_to_datetime(t64)
    timestamp_str = dt.strftime('%Y%m%d_%H%M%S')
    return f'{timestamp_str}{suffix}'


class Monitor:
    ST_SCAN = 0
    ST_CAPTURE = 1

    def __init__(self, driver, device_prefix):
        self._driver = driver
        self._device_prefix = device_prefix
        self._fs = 2_000_000 if 'JS110' in device_prefix else 1_000_000  # Hz
        self._summary_fs = 50  # Hz
        self._summary_sample_id = 0
        self._summary_sample_id_offset = 0
        self._on_statistic_fn = self._on_statistic
        self._i_info = None
        self._v_info = None
        self._end_count = 0
        self._end_count_threshold = 2 * self._summary_fs  # 2 seconds
        self._state = self.ST_SCAN
        self._i_threshold = 0.001  # 1 mA
        self._detail_jls = None
        self._summary_jls = None
        self._summary_xlsx = None
        self._summary_xlsx_filename = None
        self._stat_buffer = []

        # Configure memory buffer for full-rate sample data
        self._rsp_topic = 'm/mem/001/!rsp'
        driver.publish('m/@/!add', 1)
        driver.publish('m/001/a/!add', 1)
        driver.publish('m/001/a/!add', 2)
        driver.publish('m/001/g/size', 500_000_000)
        driver.publish('m/001/s/001/topic', f'{device_prefix}/s/i/!data')
        driver.publish('m/001/s/002/topic', f'{device_prefix}/s/v/!data')
        driver.subscribe('m/001/s/001/info', ['pub'], self._on_buffer_i_info)
        driver.subscribe('m/001/s/002/info', ['pub'], self._on_buffer_v_info)
        driver.subscribe(self._rsp_topic, ['pub'], self._on_mem_rsp)

    def _open_jls(self, filename, fs):
        """Open a JLS v2 file for the connected Joulescope.

        :param filename: The file path for the JLS write.
        :param fs: The sampling frequency in Hertz.
        :return: The open JLS file ready for writing.
        """
        _, model, serial_number = self._device_prefix.split('/')
        model = model.upper()
        wr = Writer(filename)
        wr.source_def(
            source_id=1,
            name=f'{model}-{serial_number}',
            vendor='Jetperch',
            model=model,
            version='',
            serial_number=serial_number,
        )
        wr.signal_def(
            signal_id=1,
            source_id=1,
            signal_type=SignalType.FSR,
            data_type=DataType.F32,
            sample_rate=fs,
            name='current',
            units='A',
        )
        wr.signal_def(
            signal_id=2,
            source_id=1,
            signal_type=SignalType.FSR,
            data_type=DataType.F32,
            sample_rate=fs,
            name='voltage',
            units='V',
        )
        return wr

    def _on_buffer_i_info(self, topic, value):
        # memory buffer contents for current
        self._i_info = value

    def _on_buffer_v_info(self, topic, value):
        # memory buffer contents for voltage
        self._v_info = value

    def _on_statistic(self, topic, value):
        """Callback for 50 Hz statistics."""
        # print(value)
        utc = value['time']['utc']['value'][0]
        i_avg = value['signals']['current']['avg']['value']
        v_avg = value['signals']['current']['avg']['value']
        self._stat_buffer.append((utc, i_avg, v_avg))
        self._stat_buffer = self._stat_buffer[-self._summary_fs:]
        # print(f'{i_avg} {v_avg}')
        if self._state == self.ST_SCAN:
            self._end_count = 0
            self._summary_sample_id = 0
            self._summary_sample_id_offset = 0
            if i_avg >= self._i_threshold:
                dt = _time64_to_datetime(utc).isoformat()
                print(f'{dt}: Current threshold exceeded.  Start capture')
                self._state = self.ST_CAPTURE
                summary_jls_filename = _time64_to_filename(utc, '_summary.jls')
                self._summary_jls = self._open_jls(summary_jls_filename, self._summary_fs)
                utc_v = self._stat_buffer[0][0]
                self._summary_jls.utc(1, 0, utc_v)
                self._summary_jls.utc(2, 0, utc_v)
                i_array = np.array([x[1] for x in self._stat_buffer], dtype=np.float32)
                v_array = np.array([x[2] for x in self._stat_buffer], dtype=np.float32)
                self._summary_jls.fsr_f32(1, 0, i_array)
                self._summary_jls.fsr_f32(2, 0, v_array)
                self._summary_sample_id = len(self._stat_buffer)
                self._summary_sample_id_offset = - (self._summary_sample_id - 1)

                self._summary_xlsx_filename = _time64_to_filename(utc, '_summary.xlsx')
                self._summary_xlsx = openpyxl.Workbook()
                worksheet = self._summary_xlsx.active
                worksheet.append(['SampleID', 'Current A', 'Voltage V'])
                sample_id_zero = len(self._stat_buffer) - 1
                for idx, (_, i_v, v_v) in enumerate(self._stat_buffer):
                    row = idx + 2
                    worksheet.cell(row=row, column=1, value=idx - sample_id_zero)
                    worksheet.cell(row=row, column=2, value=i_v)
                    worksheet.cell(row=row, column=3, value=v_v)
            return

        self._summary_jls.fsr_f32(1, self._summary_sample_id, np.array([i_avg], dtype=np.float32))
        self._summary_jls.fsr_f32(2, self._summary_sample_id, np.array([v_avg], dtype=np.float32))
        row = self._summary_sample_id + 2
        if self._summary_xlsx is not None:
            worksheet = self._summary_xlsx.active
            worksheet.cell(row=row, column=1, value=self._summary_sample_id + self._summary_sample_id_offset)
            worksheet.cell(row=row, column=2, value=i_avg)
            worksheet.cell(row=row, column=3, value=v_avg)

        if self._summary_sample_id > (1024 * 1024 - 8):
            self._summary_xlsx_close()

        if i_avg < self._i_threshold:
            self._end_count += 1
            if self._end_count >= self._end_count_threshold:
                self._summary_jls.utc(1, self._summary_sample_id, utc)
                self._summary_jls.utc(2, self._summary_sample_id, utc)
                dt = _time64_to_datetime(utc).isoformat()
                print(f'{dt}: Current returned to normal.  End capture')
                self._end_count = 0
                self._state = self.ST_SCAN
                self._summary_jls.close()
                self._summary_jls = None
                self._summary_xlsx_close()

        self._summary_sample_id += 1

    def _summary_xlsx_close(self):
        if self._summary_xlsx is not None:
            self._summary_xlsx.save(self._summary_xlsx_filename)
            self._summary_xlsx = None

    def _on_mem_rsp(self, topic, value):
        pass

    def device_publish(self, topic, value):
        self._driver.publish(f'{self._device_prefix}/{topic}', value)

    def open(self):
        self._driver.open(self._device_prefix)
        self.device_publish('h/fs', self._fs)
        self.device_publish('s/stats/scnt', self._fs // self._summary_fs)
        self.device_publish('s/i/range/mode', 'auto')
        self.device_publish('s/v/range/select', '15 V')
        self.device_publish('s/v/range/mode', 'manual')  # as of 2023-03-24, auto not working well
        self.device_publish('s/i/ctrl', 1)
        self.device_publish('s/v/ctrl', 1)
        self.device_publish('s/stats/ctrl', 1)
        self._driver.subscribe(f'{self._device_prefix}/s/stats/value', ['pub'], self._on_statistic_fn)

    def close(self):
        self.device_publish('s/i/ctrl', 0)
        self.device_publish('s/v/ctrl', 0)
        self.device_publish('s/stats/ctrl', 0)
        self._driver.unsubscribe(f'{self._device_prefix}/s/stats/value', self._on_statistic_fn)
        self._driver.close(self._device_prefix)
        if self._summary_jls is not None:
            self._summary_jls.close()
        self._summary_xlsx_close()


def run():
    with Driver() as driver:
        device_paths = driver.device_paths()
        if len(device_paths) == 0:
            print('No Joulescope found')
            return 1
        elif len(device_paths) > 1:
            print('Multiple Joulescope found')
            return 1
        device_prefix = device_paths[0]
        monitor = Monitor(driver, device_prefix)
        monitor.open()
        try:
            while True:
                time.sleep(0.010)
        except KeyboardInterrupt:
            pass
        monitor.close()


if __name__ == '__main__':
    run()
